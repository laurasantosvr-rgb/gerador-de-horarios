from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from datetime import timedelta

from src.models import day_off, schedule

class ScheduleExporter:
    """Exporta o horário para Excel."""

    def export(
        self,
        schedule,
        hotel,
        filename: str,
    ) -> None:

        workbook = Workbook()

        sheet = workbook.active
        sheet.title = "Horário"

        day_columns = self._create_header(
            sheet,
            schedule,
        )

        employee_rows = self._write_employees(
            sheet,
            hotel,
        )

        self._highlight_weekends(
            sheet,
            day_columns,
        )

        self._fill_schedule(
            sheet,
            schedule,
            hotel,
            employee_rows,
            day_columns,
        )

        self._format_sheet(
            sheet,
        )

        workbook.save(filename)

    def _create_header(
        self,
        sheet,
        schedule,
    ):
        """
    Cria o cabeçalho do horário.
    """

        header = [
            "Colaborador",
        ]

        days = []

        current_day = schedule.start_date

        while current_day <= schedule.end_date:

            days.append(current_day)

            current_day += timedelta(days=1)


        weekdays = [
            "Seg",
            "Ter",
            "Qua",
            "Qui",
            "Sex",
            "Sáb",
            "Dom",
        ]

        day_columns = {}

        column = 2

        for day in days:

            header.append(day)

            day_columns[day] = column

            column += 1

        sheet.append(header)

        for day, column in day_columns.items():

            cell = sheet.cell(
                row=1,
                column=column,
            )

            cell.number_format = "DD-MM-YYYY"

            if day.weekday() >= 5:

                cell.fill = PatternFill(
                    fill_type="solid",
                    start_color="D9D9D9",
                    end_color="D9D9D9",
                )

            else:

                cell.fill = PatternFill(
                    fill_type="solid",
                    start_color="E7E6E6",
                    end_color="E7E6E6",
                )

            cell.font = Font(
                bold=True,
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        sheet.row_dimensions[1].height = 25

        # Coluna "Colaborador"
        sheet["A1"].fill = PatternFill(
            fill_type="solid",
            start_color="D9D9D9",
            end_color="D9D9D9",
        )

        return day_columns

    
    def _highlight_weekends(
        self,
        sheet,
        day_columns,
    ):
        """
    Destaca as colunas de sábado e domingo.
    """

        weekend_fill = PatternFill(
            fill_type="solid",
            start_color="F2F2F2",
            end_color="F2F2F2",
        )

        for day, column in day_columns.items():

            if day.weekday() >= 5:

                for row in range(
                    1,
                    sheet.max_row + 1,
                ):

                    sheet.cell(
                        row=row,
                        column=column,
                    ).fill = weekend_fill

    def _write_employees(
        self,
        sheet,
        hotel,
    ):
        """
    Escreve os colaboradores agrupados
    pela unidade principal.

    Devolve:
        {employee_id: linha}
    """

        employee_rows = {}

        row = 2

        units = sorted(
            hotel.units.values(),
            key=lambda unit: unit.nome,
        )

        for unit in units:

            employees = sorted(
                [
                    employee
                    for employee in hotel.employees.values()
                    if employee.unidade_principal == unit.nome
                ],
                key=lambda employee: employee.nome,
            )

            for employee in employees:

                employee_rows[
                    employee.id
                ] = row

                cell = sheet.cell(
                    row=row,
                    column=1,
                )

                cell.value = employee.nome

                cell.fill = PatternFill(
                    fill_type="solid",
                    start_color="FFFFFF",
                    end_color="FFFFFF",
                )

                cell.font = Font(
                    bold=True,
                )

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

                row += 1

        return employee_rows
    
    def _unit_fill(self, unit_name):
        """
        Devolve o horário e a cor associados à unidade.
        """

        units = {
            "HM": {
                "schedule": "09:00-17:00",
                "color": "F5A92A",
            },
            "B": {
                "schedule": "09:00-17:00",
                "color": "4929D9",
            },
            "JP": {
                "schedule": "08:00-16:30",
                "color": "AD7F2F",
            },
            "AP": {
                "schedule": "11:00-15:00",
                "color": "CFF2FF",
            },
            "Lavandaria": {
                "schedule": "08:00-16:30",
                "color": "FACAEF",
            },
        }

        return units.get(
        unit_name,
            {
                "schedule": unit_name,
                "color": "FFFFFF",
            },
        )
    
    def _fill_schedule(
        self,
        sheet,
        schedule,
        hotel,
        employee_rows,
        day_columns,
    ):
        """
    Preenche o horário dos colaboradores.
    """

        print("=" * 50)
        print("Período do horário")
        print(schedule.start_date)
        print(schedule.end_date)
        print("=" * 50)

        inactive_fill = PatternFill(
            fill_type="solid",
            start_color="E02F2F",   # vermelho 
            end_color="E02F2F",
        )

        for employee in hotel.employees.values():

            if employee.ativo:
                continue

            row = employee_rows[employee.id]

            for column in day_columns.values():

                cell = sheet.cell(
                    row=row,
                    column=column,
                )

                cell.value = "Baixa"

                cell.fill = inactive_fill

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

                cell.font = Font(
                    bold=True,
                )

    # ==========================
    # Trabalho
    # ==========================
        assignments_count = {}

        for assignment in schedule.assignments:
            key = (assignment.employee_id, assignment.day)

            assignments_count[key] = assignments_count.get(key, 0) + 1

        for assignment in schedule.assignments:

            employee = hotel.employees[assignment.employee_id]

            if not employee.ativo:
                continue

            row = employee_rows[
                assignment.employee_id
            ]

            column = day_columns[
                assignment.day
            ]       

            cell = sheet.cell(
                row=row,
                column=column,
            )

            key = (
                assignment.employee_id,
                assignment.day,
            )

            info = self._unit_fill(
                assignment.unit,
            )

            if assignments_count[key] == 1:
                text = info["schedule"]
            else:
                text = assignment.unit

            if cell.value:

                cell.value += " / " + text

            else:

                cell.value = text

            if assignments_count[key] == 1:

                cell.fill = PatternFill(
                    fill_type="solid",
                    start_color=info["color"],
                    end_color=info["color"],
                )

            else:

                cell.fill = PatternFill(
                    fill_type="solid",
                    start_color="EAD7C0",   # beje
                    end_color="EAD7C0",
                )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        # ==========================
        # Folgas
        # ==========================

        for day_off in schedule.day_offs:

            employee = hotel.employees[day_off.employee_id]

            if not employee.ativo:
                continue

            row = employee_rows[
                day_off.employee_id
            ]

            print("Intervalo do horário:")
            print(min(day_columns.keys()), "->", max(day_columns.keys()))

            print("Folga:")
            print(day_off.day)
            if day_off.day not in day_columns:

                print("\n===== ERRO =====")
                print("Folga fora do intervalo:", day_off.day)
                print(
                    "Intervalo disponível:",
                    min(day_columns.keys()),
                    "->",
                    max(day_columns.keys())
                )

                raise KeyError(day_off.day)

            column = day_columns[
                day_off.day
            ]
            cell = sheet.cell(
                row=row,
                column=column,
            )

            cell.value = "F"

            cell.fill = PatternFill(
                fill_type="solid",
                start_color="B5B5B5",
                end_color="B5B5B5",
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            translation = {
                0: "segunda",
                1: "terça",
                2: "quarta",
                3: "quinta",
                4: "sexta",
                5: "sábado",
                6: "domingo",
            }

        for fixed_day in hotel.fixed_days_off:

            employee = hotel.employees[fixed_day.employee_id]

            if not employee.ativo:
                continue

            row = employee_rows[
                fixed_day.employee_id
            ]

            for day, column in day_columns.items():

                if (
                    translation[day.weekday()]
                    != fixed_day.weekday
                ):
                    continue

                cell = sheet.cell(
                    row=row,
                    column=column,
                )
                 
                cell.value = "F"

                cell.fill = PatternFill(
                    fill_type="solid",
                    start_color="B5B5B5",
                    end_color="B5B5B5",
                )

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

        for employee in hotel.employees.values():

            if employee.trabalha_fins_de_semana:
                continue

            row = employee_rows[employee.id]

            if not employee.ativo:
                continue

            for day, column in day_columns.items():

                if day.weekday() not in (5, 6):
                    continue

                cell = sheet.cell(
                    row=row,
                    column=column,
                )

                cell.value = "F"

                cell.fill = PatternFill(
                    fill_type="solid",
                    start_color="B5B5B5",
                    end_color="B5B5B5",
                )

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )   

        # ==========================
        # Férias
        # ==========================

        ferias_fill = PatternFill(
            fill_type="solid",
            start_color="FCFC44",   # amarelo
            end_color="FCFC44",
        )

        for vacation in hotel.vacations:

            employee = hotel.employees[vacation.colaborador_id]

            if not employee.ativo:
                continue

            row = employee_rows[vacation.colaborador_id]

            day = vacation.data_inicio

            while day <= vacation.data_fim:

                if day in day_columns:

                    column = day_columns[day]

                    cell = sheet.cell(
                        row=row,
                        column=column,
                    )

                    cell.value = "Férias"

                    cell.fill = ferias_fill

                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                    )

                day += timedelta(days=1)

    def _format_sheet(
        self,
        sheet,
    ):
        """
    Aplica a formatação geral da folha.
    """

        sheet.freeze_panes = "B2"

        for column in sheet.columns:

            length = max(
                len(str(cell.value))
                if cell.value is not None
                else 0
                for cell in column
            )

            sheet.column_dimensions[
                get_column_letter(column[0].column)
            ].width = length + 2

        thin = Side(
            border_style="thin",
            color="D9D9D9",
        )

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )

        for row in sheet.iter_rows():
        
            sheet.row_dimensions[
                row[0].row
            ].height = 35

            for cell in row:

                cell.border = border

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )
    
    
