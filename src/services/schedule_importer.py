from datetime import date
from openpyxl import load_workbook
from datetime import datetime

from src.models.schedule import Schedule


class ScheduleImporter:
    """
    Lê um horário em Excel e reconstrói
    apenas as folgas dos colaboradores.
    """

    def load(
        self,
        excel_path,
        hotel,
    ) -> Schedule:

        workbook = load_workbook(
            excel_path,
            data_only=True,
        )

        sheet = workbook.active

        day_columns = self._read_header(
            sheet,
        )

        start_date = min(
            day_columns.values(),
        )

        end_date = max(
            day_columns.values(),
        )

        schedule = Schedule(
            start_date=start_date,
            end_date=end_date,
        )

        employees = {
            employee.nome: employee
            for employee in hotel.employees.values()
        }

        self._read_day_offs(
            sheet,
            employees,
            schedule,
            day_columns,
        )

        return schedule

    def _read_header(
        self,
        sheet,
    ):
        """
    Lê o cabeçalho do Excel.

    Devolve:
        {coluna: data}
    """

        day_columns = {}

        column = 2

        while True:

            value = sheet.cell(
                row=1,
                column=column,
            ).value

            if value is None:
                break

            #
            # openpyxl devolve normalmente um datetime,
            # mas aceitamos também date por segurança.
            #

            if isinstance(
                value,
                datetime,
            ):
                value = value.date()

            day_columns[column] = value

            column += 1

        if not day_columns:
            raise ValueError(
                "O ficheiro do horário não contém datas no cabeçalho."
            )

        return day_columns
    
    def _read_day_offs(
        self,
        sheet,
        employees,
        schedule,
        day_columns,
    ):
        """
    Lê as folgas dos colaboradores
    a partir do Excel.
    """

        row = 2

        while True:

            employee_name = sheet.cell(
                row=row,
                column=1,
            ).value

            #
            # Fim da lista de colaboradores
            #

            if employee_name is None:
                break

            employee_name = str(
                employee_name,
            ).strip()

            employee = employees.get(
                employee_name,
            )      

            if employee is None:

                row += 1

                continue

            #
            # Percorre todos os dias
            #

            for column, day in day_columns.items():

                value = sheet.cell(
                    row=row,
                    column=column,
                ).value

                if value is None:
                    continue

                value = str(
                    value,
                ).strip()

                #
                # Apenas interessa a folga
                #

                if value == "F":

                    schedule.add_day_off(
                        employee.id,
                        day,
                    )

            row += 1